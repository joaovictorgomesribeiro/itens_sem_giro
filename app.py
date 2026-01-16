import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import shutil
from io import BytesIO

st.set_page_config(page_title="ITENS SEM GIRO", layout="centered")

st.title("📊 ITENS SEM GIRO > 90 DIAS")
st.markdown("Suba o arquivo consolidado do SAC e receba os relatórios das filiais formatados em um arquivo ZIP.")

uploaded_file = st.file_uploader("Escolha o arquivo Excel (.xlsx)", type="xlsx")

if uploaded_file:
    if st.button("Processar e Gerar Relatórios"):
        with st.spinner('Processando relatórios...'):
            # Limpeza de pastas temporárias
            pasta_saida = 'temp_relatorios'
            if os.path.exists(pasta_saida): shutil.rmtree(pasta_saida)
            os.makedirs(pasta_saida)

            # Lógica de processamento
            df_completo = pd.read_excel(uploaded_file, skiprows=1)
            novas_colunas = list(df_completo.columns)
            novas_colunas[4], novas_colunas[5], novas_colunas[6], novas_colunas[7] = \
                "Qtde de Estoque", "Valor Estoque", "Qtde de SKUs", "Dias Sem Giro"
            df_completo.columns = novas_colunas
            df_completo['Cód Empresa'] = pd.to_numeric(df_completo['Cód Empresa'], errors='coerce')

            filiais = [1, 2, 3, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 24, 27, 29, 30, 39, 40, 41]
            
            azul_escuro = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            fonte_branca = Font(color='FFFFFF', bold=True)
            negrito = Font(bold=True)
            formato_milhar = '#,##0'

            for filial in filiais:
                df_filtrado = df_completo[df_completo['Cód Empresa'] == filial].copy()
                if not df_filtrado.empty:
                    caminho = os.path.join(pasta_saida, f'{filial}.xlsx')
                    df_filtrado.to_excel(caminho, index=False)
                    
                    wb = load_workbook(caminho)
                    ws = wb.active
                    ws.title = str(filial)
                    if "Apêndice" in wb.sheetnames: del wb["Apêndice"]
                    
                    ws.insert_rows(2)
                    ws.cell(row=2, column=1).value = "RESUMO"
                    ws.cell(row=2, column=1).font = negrito
                    
                    for col in range(5, 9):
                        letra = ws.cell(row=1, column=col).column_letter
                        ws.cell(row=2, column=col).value = f"=AVERAGE({letra}3:{letra}{ws.max_row})" if letra == 'H' else f"=SUM({letra}3:{letra}{ws.max_row})"
                        ws.cell(row=2, column=col).font = negrito
                        ws.cell(row=2, column=col).number_format = formato_milhar
                    
                    for col in range(1, 9):
                        ws.cell(row=1, column=col).fill = azul_escuro
                        ws.cell(row=1, column=col).font = fonte_branca
                    
                    for row in range(3, ws.max_row + 1):
                        for col in range(5, 9):
                            ws.cell(row=row, column=col).number_format = formato_milhar
                            
                    for col in ws.columns:
                        column_letter = col[0].column_letter
                        ws.column_dimensions[column_letter].width = 20 if column_letter == 'D' else 18
                    
                    wb.save(caminho)

            # Criar o ZIP
            shutil.make_archive('Relatorios_Finais', 'zip', pasta_saida)
            
            with open("Relatorios_Finais.zip", "rb") as f:
                st.download_button(
                    label="📥 Baixar todos os relatórios (ZIP)",
                    data=f,
                    file_name="Relatorios_Finais.zip",
                    mime="application/zip"
                )
            st.success("Tudo pronto!")
